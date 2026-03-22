"""
End-to-end ML review analysis pipeline.

Steps:
  1. Score each review (sentiment, severity, actionability)
  2. Embed reviews with sentence-transformers
  3. Reduce dimensions with UMAP (384D → 5D + 2D)
  4. Cluster with HDBSCAN
  5. Compute blended confidence scores
  6. Auto-label clusters via TF-IDF
  7. Find representative review per cluster
  8. Aggregate cluster statistics
  9. Write JSON output

Output JSON schema:
  {
    "meta":             { ... pipeline config ... },
    "cluster_summary":  [ { cluster stats }, ... ],  sorted by priority_score DESC
    "reviews":          [ { per-review data }, ... ]
  }
"""

import json
from collections import defaultdict

import numpy as np

from .scoring    import score_sentiment, score_severity, score_actionability
from .embedding  import embed_texts, reduce_dimensions
from .clustering import cluster_embeddings, compute_confidence, find_representatives
from .labelling  import label_clusters
from .signals    import extract_cluster_signals
from .phrases    import extract_cluster_phrases
from .taxonomy   import (
    assign_taxonomy, assign_l1_globally, generate_cluster_taxonomy,
    load_taxonomy_history, save_taxonomy_history,
)


def run_pipeline(reviews: list[dict], output_path: str) -> dict:
    """
    Run the full clustering & scoring pipeline.

    Args:
        reviews:     List of dicts with at minimum keys: id, name, date, text.
        output_path: File path for the JSON output.

    Returns:
        The complete output dict (also written to output_path).
    """
    _sep()
    print("REVIEW CLUSTERING PIPELINE")
    _sep()

    texts = [r["text"] for r in reviews]

    # Step 1: Score
    _step(1, "Scoring reviews")
    scored = []
    for r in reviews:
        scored.append({
            **r,
            "sentiment":          score_sentiment(r["text"]),
            "severity":           score_severity(r["text"]),
            "actionability":      score_actionability(r["text"]),
            "cluster_id":         None,
            "cluster_label":      None,
            "cluster_confidence": None,
            "is_noise":           None,
            "is_representative":  None,
            "umap_2d_x":          None,
            "umap_2d_y":          None,
        })
    print(f"  Scored {len(scored)} reviews.")

    # Step 2: Embed
    _step(2, "Embedding")
    embeddings = embed_texts(texts)
    print(f"  Shape: {embeddings.shape}")

    # Step 3: Reduce
    _step(3, "UMAP dimension reduction")
    reduced_5d, coords_2d = reduce_dimensions(embeddings, n_components=5)
    print(f"  Reduced shape: {reduced_5d.shape}")

    # Step 4: Cluster
    _step(4, "HDBSCAN clustering")
    labels, hdb_probs = cluster_embeddings(reduced_5d)

    # Step 5: Confidence
    _step(5, "Confidence scoring")
    confidence = compute_confidence(reduced_5d, labels, hdb_probs)

    # Step 6: Label
    _step(6, "Auto-labelling clusters")
    cluster_info = label_clusters(reviews, labels)
    for cid, info in sorted(cluster_info.items()):
        print(f"  Cluster {cid:2d}: {info['label']:<35s}  kw={info['keywords']}")

    # Step 7: Representatives
    reps = find_representatives(reduced_5d, labels)

    # Step 8: Attach to reviews
    _step(7, "Attaching cluster data")
    for i, (review, label, conf) in enumerate(zip(scored, labels, confidence)):
        cid = int(label)
        info = cluster_info.get(cid, {})
        review.update({
            "cluster_id":         cid,
            "cluster_label":      info.get("label", f"Cluster {cid}"),
            "cluster_confidence": float(conf),
            "is_noise":           cid == -1,
            "is_representative":  reps.get(cid) == i,
            "umap_2d_x":          round(float(coords_2d[i][0]), 6),
            "umap_2d_y":          round(float(coords_2d[i][1]), 6),
        })

    # Step 8b: TF-IDF signals
    _step(8, "Extracting TF-IDF signals")
    cluster_signals = extract_cluster_signals(scored)

    # Step 8c: Repeated phrases
    _step(9, "Detecting repeated phrases")
    cluster_phrases = extract_cluster_phrases(scored)

    # Step 9: Aggregate (now includes signals, phrases, taxonomy, color)
    _step(10, "Aggregating cluster statistics")
    cluster_summary = _aggregate(
        reviews, labels, cluster_info, scored,
        cluster_signals, cluster_phrases,
    )

    output = {
        "meta": {
            "total_reviews":   len(reviews),
            "n_clusters":      sum(1 for c in cluster_summary if not c["is_noise"]),
            "n_noise_points":  int(sum(1 for l in labels if l == -1)),
            "embedding_model": "sentence-transformers/all-MiniLM-L6-v2",
            "embedding_dims":  int(embeddings.shape[1]),
            "umap_dims":       5,
            "clustering_algo": "HDBSCAN (sklearn)",
            "hdbscan_params": {
                "min_cluster_size":          2,
                "min_samples":               1,
                "cluster_selection_epsilon": 0.3,
                "cluster_selection_method":  "eom",
            },
        },
        "cluster_summary": cluster_summary,
        "reviews":         scored,
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
    print(f"\nJSON output written to: {output_path}")

    # Auto-generate companion Excel output
    excel_path = output_path.replace(".json", "_output.xlsx")
    try:
        _write_output_excel(output, excel_path)
        print(f"Excel output written to: {excel_path}")
    except Exception as e:
        print(f"  Warning: Excel output failed: {e}")

    # Summary table
    _sep()
    print("CLUSTER PRIORITY RANKING")
    _sep()
    print(f"{'#':<3} {'Label':<35} {'N':>5}  {'Priority':>8}  {'Sentiment':>9}  {'HighSev':>7}")
    print("-" * 72)
    for rank, c in enumerate(cluster_summary, 1):
        if not c["is_noise"]:
            print(
                f"{rank:<3} {c['cluster_label']:<35} {c['review_count']:>5}  "
                f"{c['priority_score']:>8.3f}  {c['avg_sentiment']:>9.3f}  "
                f"{c['high_severity_count']:>7}"
            )

    return output


# ── Helpers ───────────────────────────────────────────────────────────────────

_SEVERITY_ORDER = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}

_CLUSTER_COLORS = [
    "#f05454", "#9b70f5", "#f0a030", "#f07840",
    "#3fbe7a", "#4da8f5", "#f0c454", "#e056a0",
    "#56c8c8", "#a0a0ff", "#80c060", "#d07030",
]


def _aggregate(
    reviews: list[dict],
    labels,
    cluster_info: dict,
    scored: list[dict],
    cluster_signals: dict | None = None,
    cluster_phrases: dict | None = None,
) -> list[dict]:
    cluster_signals = cluster_signals or {}
    cluster_phrases = cluster_phrases or {}

    clusters: dict[int, list[int]] = defaultdict(list)
    for i, label in enumerate(labels):
        clusters[int(label)].append(i)

    summary = []
    color_idx = 0

    for cid, indices in clusters.items():
        cr   = [scored[i] for i in indices]
        info = cluster_info.get(cid, {})

        sent_scores = [r["sentiment"]["compound"] for r in cr]
        sev_scores  = [r["severity"]["score"] for r in cr]
        act_scores  = [r["actionability"]["score"] for r in cr]
        conf_scores = [r["cluster_confidence"] for r in cr]

        neg   = sum(1 for r in cr if r["sentiment"]["label"] == "Negative")
        pos   = sum(1 for r in cr if r["sentiment"]["label"] == "Positive")
        hi_sv = sum(1 for r in cr if r["severity"]["label"] in ("Critical", "High"))

        avg_sent = round(float(np.mean(sent_scores)), 4)
        label_str = info.get("label", f"Cluster {cid}")

        color = _CLUSTER_COLORS[color_idx % len(_CLUSTER_COLORS)] if cid != -1 else "#666666"
        if cid != -1:
            color_idx += 1

        summary.append({
            "cluster_id":            cid,
            "cluster_label":         label_str,
            "is_noise":              cid == -1,
            "top_keywords":          info.get("keywords", []),
            "review_count":          len(indices),
            "review_ids":            [reviews[i]["id"] for i in indices],
            "avg_sentiment":         avg_sent,
            "avg_severity_score":    round(float(np.mean(sev_scores)), 4),
            "avg_actionability":     round(float(np.mean(act_scores)), 4),
            "avg_confidence":        round(float(np.mean(conf_scores)), 4),
            "negative_review_count": neg,
            "positive_review_count": pos,
            "high_severity_count":   hi_sv,
            "priority_score":        round(
                0.5 * float(np.mean(sev_scores))
                + 0.3 * (neg / len(indices))
                + 0.2 * min(len(indices) / 10, 1.0),
                4,
            ),
            "signals":  cluster_signals.get(cid, []),
            "phrases":  cluster_phrases.get(cid, {}),
            "taxonomy": {},
            "color":    color,
        })

    # ── Dynamic LLM taxonomy ────────────────────────────────────────────
    non_noise = [c for c in summary if not c["is_noise"]]
    if non_noise:
        print("  Generating dynamic L1 vocabulary via LLM...")
        history = load_taxonomy_history()
        previous_vocab = None
        if history:
            try:
                last_date = max(history.keys())
                prev_snap = history[last_date]
                previous_vocab = list({
                    v.get("l1") for v in prev_snap.values() if v.get("l1")
                })
            except Exception:
                pass

        l1_map, vocab = assign_l1_globally(non_noise, previous_vocab=previous_vocab)

        # Assign taxonomy per cluster (largest first for unique themes)
        sorted_by_size = sorted(non_noise, key=lambda c: c["review_count"], reverse=True)
        used_themes = []

        for c in sorted_by_size:
            cid = c["cluster_id"]
            l1_label = l1_map.get(cid, vocab[0] if vocab else "Product Experience")
            tax = generate_cluster_taxonomy(
                c["cluster_label"],
                c["top_keywords"],
                c["avg_sentiment"],
                c["review_count"],
                l1_label,
                used_themes=used_themes,
            )
            c["taxonomy"] = tax
            theme = tax.get("theme", "")
            if theme:
                used_themes.append(theme)

        # Save taxonomy snapshot
        from datetime import date
        snap = {str(c["cluster_id"]): c["taxonomy"] for c in non_noise}
        history[date.today().isoformat()] = snap
        save_taxonomy_history(history)
    else:
        for c in summary:
            c["taxonomy"] = assign_taxonomy(c["cluster_label"], c["avg_sentiment"])

    # ── Noise cluster fallback taxonomy ─────────────────────────────────
    for c in summary:
        if c["is_noise"] and not c["taxonomy"]:
            c["taxonomy"] = assign_taxonomy(c["cluster_label"], c["avg_sentiment"])

    # ── Duplicate cluster merging ───────────────────────────────────────
    summary = _merge_duplicate_themes(summary)

    summary.sort(key=lambda x: x["priority_score"], reverse=True)
    return summary


def _merge_duplicate_themes(summary):
    """Merge clusters that ended up with identical theme labels after LLM labeling."""
    seen_themes = {}
    deduped = []

    for c in summary:
        if c["is_noise"]:
            deduped.append(c)
            continue

        theme = c.get("cluster_label", "")
        if theme in seen_themes:
            existing = deduped[seen_themes[theme]]
            n1 = existing["review_count"]
            n2 = c["review_count"]
            total = n1 + n2

            for metric in ("avg_sentiment", "avg_severity_score",
                           "avg_actionability", "avg_confidence"):
                existing[metric] = round(
                    (existing[metric] * n1 + c[metric] * n2) / total, 4
                )

            existing["review_count"] = total
            existing["review_ids"] += c["review_ids"]
            existing["negative_review_count"] += c["negative_review_count"]
            existing["positive_review_count"] += c["positive_review_count"]
            existing["high_severity_count"] += c["high_severity_count"]
            existing["priority_score"] = round(
                0.5 * existing["avg_severity_score"]
                + 0.3 * (existing["negative_review_count"] / total)
                + 0.2 * min(total / 10, 1.0),
                4,
            )

            # Merge signals and phrases
            existing_terms = {s["term"] for s in existing.get("signals", [])}
            for s in c.get("signals", []):
                if s["term"] not in existing_terms:
                    existing["signals"].append(s)
            for phrase, count in c.get("phrases", {}).items():
                existing["phrases"][phrase] = existing["phrases"].get(phrase, 0) + count

            print(f"  Merged duplicate theme '{theme}' - combined {total} reviews")
        else:
            seen_themes[theme] = len(deduped)
            deduped.append(c)

    return deduped


def _write_output_excel(output, path):
    """Generate a 4-sheet Excel workbook from pipeline output."""
    from openpyxl import Workbook

    wb = Workbook()
    reviews = output.get("reviews", [])
    clusters = output.get("cluster_summary", [])

    # Sheet 1: Reviews
    ws1 = wb.active
    ws1.title = "Reviews"
    headers = ["ID", "Name", "Date", "Review Text", "Sentiment", "Sent Score",
               "Severity", "Sev Score", "Actionability", "Act Score", "Cluster", "L1", "L2"]
    ws1.append(headers)
    for r in reviews:
        sent = r.get("sentiment", {})
        sev = r.get("severity", {})
        act = r.get("actionability", {})
        ws1.append([
            r.get("id", ""), r.get("name", ""), r.get("date", ""),
            r.get("text", ""),
            sent.get("label", ""), round(sent.get("compound", 0), 3),
            sev.get("label", ""), round(sev.get("score", 0), 3),
            act.get("label", ""), round(act.get("score", 0), 3),
            r.get("cluster_label", ""),
            r.get("cluster_label", ""),
            r.get("cluster_label", ""),
        ])

    # Sheet 2: Cluster Summary
    ws2 = wb.create_sheet("Cluster Summary")
    ws2.append(["Cluster", "Reviews", "Priority /100", "Avg Sentiment",
                "High Severity", "Avg Actionability", "Top Keywords", "Repeated Phrases"])
    for c in clusters:
        if c.get("is_noise"):
            continue
        kw = ", ".join(c.get("top_keywords", [])[:5])
        ph = ", ".join(f'"{p}"' for p in list(c.get("phrases", {}).keys())[:4])
        ws2.append([
            c.get("cluster_label", ""), c.get("review_count", 0),
            round(c.get("priority_score", 0) * 100),
            round(c.get("avg_sentiment", 0), 3),
            c.get("high_severity_count", 0),
            round(c.get("avg_actionability", 0) * 100),
            kw, ph,
        ])

    # Sheet 3: Taxonomy Tree
    ws3 = wb.create_sheet("Taxonomy Tree")
    ws3.append(["L1 Area", "L2 Feature", "L3 Specific", "Theme", "Subtheme", "Reviews", "Priority"])
    for c in clusters:
        if c.get("is_noise"):
            continue
        tax = c.get("taxonomy", {})
        ws3.append([
            tax.get("l1", ""), tax.get("l2", ""), tax.get("l3", ""),
            tax.get("theme", ""), tax.get("subtheme", ""),
            c.get("review_count", 0), round(c.get("priority_score", 0) * 100),
        ])

    # Sheet 4: CX Metrics
    ws4 = wb.create_sheet("CX Metrics")
    total = len(reviews)
    if total:
        pos_count = sum(1 for r in reviews if r.get("sentiment", {}).get("label") == "Positive")
        neg_count = sum(1 for r in reviews if r.get("sentiment", {}).get("label") == "Negative")
        avg_sent = sum(r.get("sentiment", {}).get("compound", 0) for r in reviews) / total
        promoters = sum(1 for r in reviews
                        if r.get("sentiment", {}).get("compound", 0) >= 0.5
                        and r.get("severity", {}).get("score", 0) < 0.2)
        detractors = sum(1 for r in reviews
                         if r.get("sentiment", {}).get("compound", 0) <= -0.3
                         or r.get("severity", {}).get("label") in ("Critical", "High"))
        csat = round(pos_count / total * 100, 1)
        nps = round((promoters - detractors) / total * 100, 1)
    else:
        pos_count = neg_count = promoters = detractors = 0
        avg_sent = csat = nps = 0

    ws4.append(["Metric", "Value", "Formula", "Note"])
    ws4.append(["CSAT", f"{csat}%", "% Positive reviews",
                "Good" if csat >= 50 else "Needs attention"])
    ws4.append(["NPS Proxy", f"{'+' if nps > 0 else ''}{nps}",
                "(Promoters-Detractors)/Total x 100",
                "Positive" if nps > 0 else "Negative NPS"])
    ws4.append(["Total Reviews", total, "len(reviews)", ""])
    ws4.append(["Positive Reviews", pos_count, "sentiment=Positive",
                f"{round(pos_count/max(total,1)*100)}%"])
    ws4.append(["Negative Reviews", neg_count, "sentiment=Negative",
                f"{round(neg_count/max(total,1)*100)}%"])
    ws4.append(["Promoters", promoters, "sent>=0.5 AND sev<0.2", ""])
    ws4.append(["Detractors", detractors, "sent<=-0.3 OR sev Critical/High", ""])
    ws4.append(["Clusters", len([c for c in clusters if not c.get("is_noise")]),
                "non-noise clusters", ""])

    wb.save(path)


def _sep():
    print("\n" + "=" * 60)


def _step(n: int, name: str):
    print(f"\n-- Step {n}: {name} --")
