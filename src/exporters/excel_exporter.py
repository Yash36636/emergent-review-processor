"""
Excel exporter — produces a fully-formatted, multi-sheet analysis workbook
from the pipeline output dict.

Sheets produced:
  1. All Reviews       — every review with all ML scores inline
  2. Cluster Summary   — one row per cluster, sorted by priority
  3. Cluster Deep-Dives — reviews grouped under their cluster banner
  4. Priority Dashboard — PM-ready action board
"""

from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Palette ───────────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", start_color="1F3864", end_color="1F3864")
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
SUBHDR_FILL  = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
SUBHDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
BODY_FONT    = Font(name="Calibri", size=10)
BOLD_FONT    = Font(name="Calibri", size=10, bold=True)
WRAP         = Alignment(wrap_text=True, vertical="top")
CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
TOP_CENTER   = Alignment(horizontal="center", vertical="top", wrap_text=True)
BORDER       = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

SENTIMENT_FILLS = {
    "Positive": PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE"),
    "Neutral":  PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C"),
    "Negative": PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE"),
}
SEVERITY_COLORS = {
    "Critical": "C00000",
    "High":     "E26B0A",
    "Medium":   "C09000",
    "Low":      "375623",
}

# Rotating cluster palette (up to 12 clusters)
_CLUSTER_PALETTE = [
    "FFC7CE", "FFEB9C", "F4CCFF", "FCE4D6",
    "C6EFCE", "DDEBF7", "D9E1F2", "FFF2CC",
    "E2EFDA", "FCE4D6", "DAEEF3", "F2F2F2",
]


def _cluster_fill(cluster_id: int) -> PatternFill:
    if cluster_id == -1:
        color = "F2F2F2"
    else:
        color = _CLUSTER_PALETTE[cluster_id % len(_CLUSTER_PALETTE)]
    return PatternFill("solid", start_color=color, end_color=color)


# ── Style helpers ─────────────────────────────────────────────────────────────

def _cell(ws, row, col, value, font=None, fill=None, align=None, border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = font  or BODY_FONT
    c.alignment = align or WRAP
    if fill:
        c.fill = fill
    if border:
        c.border = BORDER
    return c


def _title_row(ws, text, n_cols, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=13)
    c.fill      = HEADER_FILL
    c.alignment = CENTER
    c.border    = BORDER
    ws.row_dimensions[row].height = 32
    return row + 1


def _header_row(ws, headers, row, widths=None):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font      = SUBHDR_FONT
        c.fill      = SUBHDR_FILL
        c.alignment = CENTER
        c.border    = BORDER
    if widths:
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[row].height = 24
    ws.freeze_panes = ws.cell(row=row + 1, column=1)
    return row + 1


# ── Sheet 1: All Reviews ──────────────────────────────────────────────────────

def _sheet_all_reviews(wb: Workbook, reviews: list[dict]):
    ws = wb.active
    ws.title = "All Reviews"

    headers = [
        "#", "Reviewer", "Date", "Review Text",
        "Sentiment", "Sent. Score",
        "Severity", "Sev. Score",
        "Actionability", "Act. Score",
        "Cluster", "Confidence",
    ]
    widths = [4, 20, 13, 58, 12, 10, 10, 9, 16, 9, 30, 11]

    row = _title_row(ws, "Full Review Analysis", len(headers))
    row = _header_row(ws, headers, row, widths)

    for i, r in enumerate(reviews):
        sent   = r["sentiment"]
        sev    = r["severity"]
        act    = r["actionability"]
        cid    = r.get("cluster_id", -1)
        sfill  = SENTIMENT_FILLS.get(sent["label"])
        cfill  = _cluster_fill(cid if cid is not None else -1)

        vals = [
            r["id"],
            r["name"],
            r["date"],
            r["text"],
            sent["label"],
            sent["compound"],
            sev["label"],
            sev["score"],
            act["label"],
            act["score"],
            r.get("cluster_label", "—"),
            round(r.get("cluster_confidence") or 0, 2),
        ]

        for col, val in enumerate(vals, 1):
            fill  = None
            font  = BODY_FONT
            align = WRAP

            if col in (1, 3, 6, 8, 10, 12):
                align = TOP_CENTER
            if col == 5:    # Sentiment label
                fill = sfill
            if col == 7:    # Severity label
                color = SEVERITY_COLORS.get(sev["label"], "000000")
                font  = Font(name="Calibri", size=10, bold=True, color=color)
            if col == 11:   # Cluster
                fill = cfill

            _cell(ws, row, col, val, font=font, fill=fill, align=align)

        # Row height based on review length
        ws.row_dimensions[row].height = max(18, min(len(r["text"]) // 60 * 15, 120))
        row += 1


# ── Sheet 2: Cluster Summary ──────────────────────────────────────────────────

def _sheet_cluster_summary(wb: Workbook, cluster_summary: list[dict]):
    ws = wb.create_sheet("Cluster Summary")

    headers = [
        "Cluster", "Reviews", "% of Total",
        "Avg Sentiment", "Avg Severity", "Avg Actionability",
        "Neg Reviews", "High/Crit Sev", "Priority Score",
        "Top Keywords",
    ]
    widths = [32, 10, 12, 14, 13, 17, 12, 14, 14, 50]

    total = sum(c["review_count"] for c in cluster_summary)
    row   = _title_row(ws, "Cluster Summary — Key Themes & Priority Ranking", len(headers))
    row   = _header_row(ws, headers, row, widths)

    for rank, c in enumerate(cluster_summary, 1):
        cid   = c["cluster_id"]
        cfill = _cluster_fill(cid)
        pct   = c["review_count"] / total * 100 if total else 0

        vals = [
            c["cluster_label"],
            c["review_count"],
            f"{pct:.1f}%",
            f"{c['avg_sentiment']:+.3f}",
            f"{c['avg_severity_score']:.3f}",
            f"{c['avg_actionability']:.3f}",
            c["negative_review_count"],
            c["high_severity_count"],
            f"{c['priority_score']:.3f}",
            ", ".join(c.get("top_keywords", [])),
        ]

        for col, val in enumerate(vals, 1):
            fill  = cfill if col == 1 else None
            font  = BOLD_FONT if col == 1 else BODY_FONT
            align = WRAP if col in (1, 10) else TOP_CENTER

            # Colour priority score by value
            if col == 9:
                score = c["priority_score"]
                color = "C00000" if score >= 0.6 else "E26B0A" if score >= 0.4 else "375623"
                font  = Font(name="Calibri", size=10, bold=True, color=color)

            _cell(ws, row, col, val, font=font, fill=fill, align=align)

        ws.row_dimensions[row].height = 22
        row += 1


# ── Sheet 3: Cluster Deep-Dives ───────────────────────────────────────────────

def _sheet_deep_dives(wb: Workbook, reviews: list[dict], cluster_summary: list[dict]):
    ws = wb.create_sheet("Cluster Deep-Dives")

    col_widths = [22, 13, 58, 12, 10, 16]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    row = _title_row(ws, "Cluster Deep-Dives — Reviews Grouped by Theme", 6)
    ws.freeze_panes = "A2"

    # Group reviews by cluster_id
    groups: dict[int, list[dict]] = defaultdict(list)
    for r in reviews:
        groups[r.get("cluster_id", -1)].append(r)

    # Follow cluster_summary order (already sorted by priority)
    for c in cluster_summary:
        cid   = c["cluster_id"]
        revs  = groups.get(cid, [])
        cfill = _cluster_fill(cid)

        # Cluster banner
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        banner = ws.cell(row=row, column=1,
                         value=f"  {c['cluster_label']}  ({len(revs)} reviews)  "
                               f"|  Priority: {c['priority_score']:.3f}  "
                               f"|  Keywords: {', '.join(c.get('top_keywords', []))}")
        banner.font      = Font(name="Calibri", size=11, bold=True)
        banner.fill      = cfill
        banner.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        banner.border    = BORDER
        ws.row_dimensions[row].height = 22
        row += 1

        # Sub-headers
        sub_hdrs = ["Reviewer", "Date", "Review Text", "Sentiment", "Severity", "Actionability"]
        for col, h in enumerate(sub_hdrs, 1):
            c_ = ws.cell(row=row, column=col, value=h)
            c_.font = SUBHDR_FONT; c_.fill = SUBHDR_FILL
            c_.alignment = CENTER; c_.border = BORDER
        ws.row_dimensions[row].height = 20
        row += 1

        # Review rows
        for r in revs:
            sent  = r["sentiment"]
            sev   = r["severity"]
            act   = r["actionability"]
            is_rep = r.get("is_representative", False)
            rfont  = Font(name="Calibri", size=10, bold=is_rep)

            row_vals = [
                r["name"], r["date"], r["text"],
                sent["label"], sev["label"], act["label"],
            ]
            for col, val in enumerate(row_vals, 1):
                fill  = None
                font  = rfont
                align = WRAP if col == 3 else TOP_CENTER

                if col == 4:
                    fill = SENTIMENT_FILLS.get(sent["label"])
                if col == 5:
                    color = SEVERITY_COLORS.get(sev["label"], "000000")
                    font  = Font(name="Calibri", size=10, bold=is_rep, color=color)

                _cell(ws, row, col, val, font=font, fill=fill, align=align)

            ws.row_dimensions[row].height = max(18, min(len(r["text"]) // 60 * 15, 100))
            row += 1

        row += 1  # spacer between clusters


# ── Sheet 4: Priority Dashboard ───────────────────────────────────────────────

_PM_ACTIONS = {
    "Billing & Credit Issues": (
        "Credits drain fast, lost on reload, double-charged. Users feel scammed.",
        "1) Show credit cost estimate BEFORE each task\n"
        "2) Fix credit-loss-on-reload bug\n"
        "3) Add spending cap / alert\n"
        "4) Transparent billing history\n"
        "5) Flexible pricing tiers",
    ),
    "App Bugs & Crashes": (
        "Blank screen after payment, Play Store deployment shows success but app never appears.",
        "1) Fix blank screen post-payment bug urgently\n"
        "2) Validate actual Play Store deployment status\n"
        "3) Add project auto-save before recharge prompts",
    ),
    "AI Quality Issues": (
        "AI misunderstands queries, forgetful, no persistent chat history.",
        "1) Implement persistent conversation memory\n"
        "2) Tune for task-specific prompts\n"
        "3) Add thumbs up/down feedback loop",
    ),
    "Misleading Free Tier": (
        "App advertised as free but requires credits immediately; Expo dependency not disclosed.",
        "1) Rewrite store listing to reflect credit model\n"
        "2) Show onboarding screen explaining credits\n"
        "3) Disclose third-party dependencies upfront",
    ),
    "Positive Experience": (
        "Segment of users genuinely love the product — especially vibe coding.",
        "1) Survey happy users to replicate success\n"
        "2) Create case studies for marketing\n"
        "3) Build referral programme",
    ),
}

_PRIORITY_COLORS = ["C00000", "E26B0A", "C09000", "375623", "2E75B6", "7030A0"]


def _sheet_dashboard(wb: Workbook, cluster_summary: list[dict], reviews: list[dict]):
    ws = wb.create_sheet("Priority Dashboard")

    headers = ["#", "Priority Score", "Cluster Theme", "Problem Statement",
               "Evidence", "Recommended Actions", "Representative Review"]
    widths  = [4, 14, 28, 45, 12, 45, 55]

    row = _title_row(ws, "PM Priority Dashboard — Action Board", len(headers))
    row = _header_row(ws, headers, row, widths)

    # Build lookup: cluster_label → representative review text
    rep_lookup: dict[str, str] = {}
    for r in reviews:
        if r.get("is_representative"):
            rep_lookup[r.get("cluster_label", "")] = f'"{r["name"]}" ({r["date"]}): {r["text"][:200]}...'

    non_noise = [c for c in cluster_summary if not c["is_noise"]]

    for rank, c in enumerate(non_noise, 1):
        label  = c["cluster_label"]
        score  = c["priority_score"]
        pcolor = _PRIORITY_COLORS[min(rank - 1, len(_PRIORITY_COLORS) - 1)]
        cfill  = _cluster_fill(c["cluster_id"])

        # Match known PM actions or use generic
        problem, action = _PM_ACTIONS.get(
            label,
            (
                f"{c['negative_review_count']} negative reviews in this cluster.",
                "Investigate top reviews and form action items.",
            ),
        )

        evidence = (
            f"{c['review_count']} reviews\n"
            f"{c['negative_review_count']} negative\n"
            f"{c['high_severity_count']} high/crit sev"
        )

        rep_text = rep_lookup.get(label, "—")

        row_vals = [rank, score, label, problem, evidence, action, rep_text]

        for col, val in enumerate(row_vals, 1):
            fill  = cfill if col == 3 else None
            font  = BODY_FONT
            align = WRAP if col >= 4 else TOP_CENTER

            if col == 1:
                font  = Font(name="Calibri", size=11, bold=True, color=pcolor)
                align = TOP_CENTER
            if col == 2:
                font  = Font(name="Calibri", size=10, bold=True, color=pcolor)

            _cell(ws, row, col, val, font=font, fill=fill, align=align)

        ws.row_dimensions[row].height = 90
        row += 1


# ── Public entry-point ────────────────────────────────────────────────────────

def create_analysis_workbook(pipeline_output: dict, output_path: str) -> None:
    """
    Build the full 4-sheet analysis workbook from `run_pipeline()` output
    and save it to `output_path`.

    Args:
        pipeline_output: Dict returned by run_pipeline() with keys:
                         "meta", "cluster_summary", "reviews".
        output_path:     Destination .xlsx file path.
    """
    reviews         = pipeline_output["reviews"]
    cluster_summary = pipeline_output["cluster_summary"]

    wb = Workbook()

    _sheet_all_reviews(wb, reviews)
    _sheet_cluster_summary(wb, cluster_summary)
    _sheet_deep_dives(wb, reviews, cluster_summary)
    _sheet_dashboard(wb, cluster_summary, reviews)

    wb.save(output_path)
    print(f"Excel workbook saved to: {output_path}")
    print(f"  {len(reviews)} reviews across {len([c for c in cluster_summary if not c['is_noise']])} clusters")
    print(f"  Sheets: All Reviews | Cluster Summary | Cluster Deep-Dives | Priority Dashboard")
