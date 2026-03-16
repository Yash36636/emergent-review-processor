"""
Parses an .xlsx file containing app reviews.

Supports two layouts:
  1. Pipeline-exported workbook (sheet "All Reviews" with columns
     Reviewer / Date / Review Text)
  2. Any generic sheet with columns that look like name, date, and text

Returns a list of (name, date, review_text) tuples — the same format as
the .docx parser, so both are interchangeable downstream.
"""

from pathlib import Path
from openpyxl import load_workbook


_NAME_HINTS = {"reviewer", "name", "user", "author", "username", "user_name"}
_DATE_HINTS = {"date", "time", "timestamp", "review_date", "created", "posted"}
_TEXT_HINTS = {"review text", "review_text", "text", "review", "comment",
               "feedback", "body", "content", "description", "message"}


def _find_column(headers: list[str], hints: set[str]) -> int | None:
    lower = [h.lower().strip() for h in headers]
    for i, h in enumerate(lower):
        if h in hints:
            return i
    for i, h in enumerate(lower):
        if any(hint in h for hint in hints):
            return i
    return None


def parse_xlsx_reviews(xlsx_path: str) -> list[tuple[str, str, str]]:
    """
    Extract reviews from an Excel workbook.

    Tries the "All Reviews" sheet first (pipeline output format), then
    falls back to the first sheet. Auto-detects name / date / text columns
    from the header row.

    Returns:
        List of (name, date, review_text) tuples.
    """
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")

    wb = load_workbook(str(path), read_only=True, data_only=True)

    ws = wb["All Reviews"] if "All Reviews" in wb.sheetnames else wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return []

    raw_headers = [str(c) if c else "" for c in rows[0]]

    if not any(h.strip() for h in raw_headers):
        raw_headers = [str(c) if c else "" for c in rows[1]]
        data_rows = rows[2:]
    else:
        data_rows = rows[1:]

    name_idx = _find_column(raw_headers, _NAME_HINTS)
    date_idx = _find_column(raw_headers, _DATE_HINTS)
    text_idx = _find_column(raw_headers, _TEXT_HINTS)

    if text_idx is None:
        for i, h in enumerate(raw_headers):
            if h.lower().strip() not in {"#", "id", "index", "row"} and i != name_idx and i != date_idx:
                text_idx = i
                break

    if text_idx is None:
        return []

    reviews = []
    for row in data_rows:
        if row is None:
            continue
        text = str(row[text_idx]).strip() if text_idx < len(row) and row[text_idx] else ""
        if not text or text.lower() in ("none", "nan", "null", ""):
            continue

        name = str(row[name_idx]).strip() if name_idx is not None and name_idx < len(row) and row[name_idx] else "Unknown"
        date = str(row[date_idx]).strip() if date_idx is not None and date_idx < len(row) and row[date_idx] else "Unknown"
        if name.lower() in ("none", "nan", "null"):
            name = "Unknown"
        if date.lower() in ("none", "nan", "null"):
            date = "Unknown"

        reviews.append((name, date, text))

    return reviews
