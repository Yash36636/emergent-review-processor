"""
Voice of Customer Intelligence Platform - Pipeline v4
======================================================
Method: BERTopic Hierarchical Topic Modelling + KeyBERT + LLM Labelling

Install:
    pip install bertopic keybert sentence-transformers hdbscan umap-learn
                vaderSentiment scikit-learn numpy anthropic openpyxl pandas

Run:
    python review_pipeline_v4.py
    python review_pipeline_v4.py --input reviews.xlsx
    python review_pipeline_v4.py --input reviews.xlsx --output analysis.json
"""

import json
import re
import os
import argparse
import hashlib
from datetime import datetime
from collections import defaultdict, Counter

import numpy as np


RAW_REVIEWS = [
    {"id":1, "name":"Sheebu Decode", "date":"2026-03-12","text":"badhiya hai perfect nice work I tested thit thak hai improvement honi chahiye aur like lag feel Ho Raha hai aur aur app mein Jo LLM jo hai acche se Hinglish mein reply badhiya kar Raha hai bas thoda bahut improvement ki jarurat hai baaki Toh bahut badhiya nice work Emergent team badhiya shuruaat hai aage Aur bhi accha karenge"},
    {"id":2, "name":"Tony Minides", "date":"2026-03-11","text":"Started ok but support has disappeared and costs spiralled through the roof. I no longer get help and it's eating $500 a day when previously this would have lasted me a week. I'm looking for alternatives"},
    {"id":3, "name":"Eddie Ledington", "date":"2026-03-11","text":"was only able to use it once until I buy the credits and im not willing to buy the credits when it says free use"},
    {"id":4, "name":"Jeremy Holland", "date":"2026-03-11","text":"Emergent has enhanced my career as a Prompt Engineer and I'm grateful to eternally grow with this God Given Company. May We Grow And Prosper Together as One"},
    {"id":5, "name":"User 5", "date":"2026-03-11","text":"very bad don't recommend. don't use soulless AI instead learn and code games yourself"},
    {"id":6, "name":"Misri Kutubuddin", "date":"2026-03-11","text":"ok"},
    {"id":7, "name":"sachin jagtap", "date":"2026-03-11","text":"eats credit very fast"},
    {"id":8, "name":"Kaylee Marie Askew", "date":"2026-03-11","text":"The search results coming up on google are kind of confusing compared the app is there a way to make sure I have the right contact number?"},
    {"id":9, "name":"Surajit Mondal", "date":"2026-03-11","text":"I purchased a subscription for the Emergent app, but the service quality has not improved at all. The AI frequently misunderstands my questions and gives irrelevant answers. For a paid service, I expected much better performance and support. Currently, it feels like the subscription is not worth the money. Please improve the AI accuracy and user experience. The app scammed people, don't use this app. They do not even reply to any review."},
    {"id":10, "name":"Sumeet Kumar Gupta", "date":"2026-03-10","text":"this app is not good. please don't trust the good reviews, I'm stuck!!!"},
    {"id":11, "name":"KJ Snyder", "date":"2026-03-10","text":"Lost all purchased credits from subscription upon reloading app from another screen. I lost more than 100 credits for absolutely nothing and the app won't even reload my credits."},
    {"id":12, "name":"harsha koti", "date":"2026-03-10","text":"changes requested even because it was charged earlier are charged again. just losing money over and over again."},
    {"id":13, "name":"Praveen Singh", "date":"2026-03-09","text":"i didn't feel good to use it"},
    {"id":14, "name":"Oluseyi Oyetunde", "date":"2026-03-09","text":"awesome"},
    {"id":15, "name":"Barry Mitcheson", "date":"2026-03-09","text":"spent all day creating an app, had to recharge credit 3 times, lots of glitches, ended up with a great app, had to recharge to deploy app, AS SOON AS I PAYED FOR RECHARGE IT ALL DISAPPEARED Then NOTHING! APP OPENS UP TO A BLANK SCREEN, NOT ONLY HAVE THEY SCAMMED ME, BUT NOW THEY HAVE A UNIQUE APP WHICH THEY WILL LAUNCH FOR THEMSELVES!"},
    {"id":16, "name":"Christina", "date":"2026-03-09","text":"I enjoy this app the pricing is spendy I go thru on there too fast..."},
    {"id":17, "name":"jayshil jayshil", "date":"2026-03-09","text":"very bad experience"},
    {"id":18, "name":"Hitesh Sawlani", "date":"2026-03-08","text":"ok not good"},
    {"id":19, "name":"DEEPAK DUBBEY", "date":"2026-03-08","text":"nice"},
    {"id":20, "name":"VIGNESH SHETTY", "date":"2026-03-08","text":"I created a app turns out the app didn't got deployed in google playstore however emergent ai is showing its deployed"},
    {"id":21, "name":"Evyavan Acharya", "date":"2026-03-08","text":"after I created my first project and after I completed the creation it shows Credit exosted buy credit and it also requires expo named app it is not free at all!"},
    {"id":22, "name":"Saugat", "date":"2026-03-08","text":"it sucks"},
    {"id":23, "name":"giovany z", "date":"2026-03-08","text":"don't waste money on garbage. the agent besides being slower than digging holes, is also super forgetful. no chat history, no accountability. it's working lots and sucks at what it does."},
    {"id":24, "name":"Raymond Kirschenman", "date":"2026-03-07","text":"Gave a sponsorship to Ben Esherik telling him to be controversial. Also this genuinely just doesn't work, most of the code flaws at simple security. Don't use this, the company and app is garbage."},
    {"id":25, "name":"Ricky King", "date":"2026-03-07","text":"love the app, I do think the credits cost a bit much but I'd say worth it."},
    {"id":26, "name":"Bruno Costa", "date":"2026-03-07","text":"great experience. great vibe coding tool."},
    {"id":27, "name":"Umason Sonar", "date":"2026-03-07","text":"vary bad"},
    {"id":28, "name":"Aysha Ali", "date":"2026-03-07","text":"bakwaas app hai ye"},
    {"id":29, "name":"Makbull Patil", "date":"2026-03-07","text":"great app very nice"},
]


def load_from_excel(path: str) -> list[dict]:
    import pandas as pd
    df = pd.read_excel(path, dtype=str)
    df.columns = [c.strip().lower() for c in df.columns]

    def find_col(aliases):
        for a in aliases:
            if a in df.columns:
                return a
        for col in df.columns:
            for a in aliases:
                if a in col:
                    return col
        return None

    name_col = find_col(["name","reviewer","author","user","customer"])
    date_col = find_col(["date","reviewed_on","review_date","submitted","created_at"])
    text_col = find_col(["text","review","feedback","comment","body","message"])

    if not text_col:
        raise ValueError(
            f"No text column found. Columns in file: {list(df.columns)}\n"
            f"Rename your review column to one of: text, review, feedback, comment"
        )

    reviews = []
    for idx, row in df.iterrows():
        text = str(row.get(text_col, "")).strip()
        if not text or text.lower() in ("nan","none",""):
            continue
        reviews.append({
            "id": idx + 1,
            "name": str(row[name_col]).strip() if name_col else f"User {idx+1}",
            "date": str(row[date_col]).strip() if date_col else "",
            "text": text,
        })
    print(f"  Loaded {len(reviews)} reviews from {path}")
    return reviews


def embed_reviews(texts: list[str]) -> tuple:
    from sentence_transformers import SentenceTransformer
    print("  Loading all-MiniLM-L6-v2...")
    model = SentenceTransformer("all-MiniLM-L6-v2")
    embeddings = model.encode(texts, batch_size=32, show_progress_bar=True, normalize_embeddings=True)
    print(f"  Embeddings shape: {embeddings.shape}")
    return embeddings, model


def _kmeans_fallback(embeddings: np.ndarray, k: int = 6) -> np.ndarray:
    from sklearn.cluster import KMeans
    km = KMeans(n_clusters=k, random_state=42, n_init=10)
    labels = km.fit_predict(embeddings)
    print(f"  KMeans fallback: {k} clusters for {len(labels)} reviews")
    return labels.astype(int)


def build_topic_hierarchy(texts: list[str], embeddings: np.ndarray) -> tuple:
    from bertopic import BERTopic
    from umap import UMAP
    import hdbscan

    n = len(texts)
    n_neighbors = max(2, min(5, n // 6))
    target_topics = max(5, min(15, n // 4))

    umap_model = UMAP(n_components=min(5, n - 2), n_neighbors=n_neighbors,
                      min_dist=0.0, metric="cosine", random_state=42)
    hdbscan_model = hdbscan.HDBSCAN(min_cluster_size=2, min_samples=1,
                                     cluster_selection_epsilon=0.0, metric="euclidean",
                                     cluster_selection_method="leaf", prediction_data=True)
    topic_model = BERTopic(umap_model=umap_model, hdbscan_model=hdbscan_model,
                           nr_topics=target_topics, calculate_probabilities=False, verbose=False)
    doc_topics, _ = topic_model.fit_transform(texts, embeddings=embeddings)

    n_topics = len(set(t for t in doc_topics if t != -1))
    n_outliers = sum(1 for t in doc_topics if t == -1)
    print(f"  BERTopic found {n_topics} topics, {n_outliers} outliers (target {target_topics})")

    MIN_CLUSTERS = 5
    if n_topics < MIN_CLUSTERS:
        k = max(MIN_CLUSTERS, min(10, n // 3))
        print(f"  Only {n_topics} topics, need >= {MIN_CLUSTERS}. Running KMeans (k={k}).")
        doc_topics = _kmeans_fallback(embeddings, k=k)
        n_topics = len(set(t for t in doc_topics if t >= 0))

    hierarchy_df = topic_model.hierarchical_topics(texts)
    flat_topics = {}
    topic_embeddings = getattr(topic_model, "topic_embeddings_", None)
    for tid in set(t for t in doc_topics if t != -1):
        doc_indices = [i for i, t in enumerate(doc_topics) if t == tid]
        centroid = (topic_embeddings[tid]
                    if (topic_embeddings is not None and tid < len(topic_embeddings))
                    else embeddings[doc_indices].mean(axis=0))
        flat_topics[tid] = {"docs": doc_indices, "centroid": centroid}
    topic_tree = _extract_three_level_tree(hierarchy_df, flat_topics)
    return topic_model, doc_topics, hierarchy_df, topic_tree, flat_topics


def _extract_three_level_tree(hierarchy_df, flat_topics: dict) -> dict:
    if hierarchy_df is None or len(hierarchy_df) == 0:
        return {tid: {"l1_id": 0, "l2_id": tid, "l3_id": tid, "depth": 1,
                      "ancestry_chain": [tid]} for tid in flat_topics}

    parent_of = {}
    children_of = defaultdict(list)
    merge_dist = {}
    for _, row in hierarchy_df.iterrows():
        parent_id = int(row["Parent_ID"])
        dist = float(row.get("Distance", row.get("distance", 0.0)))
        merge_dist[parent_id] = dist
        for child_key in ["Child_Left_ID", "Child_Right_ID"]:
            if child_key in row:
                child_id = int(row[child_key])
                parent_of[child_id] = parent_id
                children_of[parent_id].append(child_id)

    distances = sorted(merge_dist.values())
    if len(distances) >= 3:
        l1_threshold = np.percentile(distances, 70)
    elif len(distances) == 2:
        l1_threshold = distances[-1]
    else:
        l1_threshold = distances[0] if distances else 1.0

    def get_ancestors(tid):
        chain = [tid]
        cur = tid
        while cur in parent_of:
            cur = parent_of[cur]
            chain.append(cur)
        return chain

    topic_tree = {}
    for tid in flat_topics:
        chain = get_ancestors(tid)
        l1_id, l3_id, l2_id = chain[-1], tid, tid
        for ancestor in reversed(chain[1:]):
            d = merge_dist.get(ancestor, 0.0)
            if d < l1_threshold:
                l2_id = ancestor
                break
        topic_tree[tid] = {"l1_id": l1_id, "l2_id": l2_id, "l3_id": l3_id,
                           "depth": len(chain), "ancestry_chain": chain}

    unique_l1 = len(set(v["l1_id"] for v in topic_tree.values()))
    unique_l2 = len(set(v["l2_id"] for v in topic_tree.values()))
    if unique_l1 == 1 and len(flat_topics) >= 3:
        tids_sorted = sorted(flat_topics.keys())
        mid = len(tids_sorted) // 2
        for i, tid in enumerate(tids_sorted):
            topic_tree[tid]["l1_id"] = 0 if i < mid else 1
    if unique_l2 == 1 and len(flat_topics) >= 4:
        for tid in flat_topics:
            topic_tree[tid]["l2_id"] = tid // 2
    return topic_tree


def extract_node_keywords(cluster_texts: list[str], cluster_centroid: np.ndarray,
                          st_model, top_n: int = 8) -> list[str]:
    try:
        from keybert import KeyBERT
        kw_model = KeyBERT(model=st_model)
        combined = " ".join(cluster_texts)
        keywords = kw_model.extract_keywords(combined, keyphrase_ngram_range=(1, 2),
                                             stop_words="english", top_n=top_n, diversity=0.5)
        return [kw for kw, score in keywords]
    except Exception:
        return _tfidf_keywords_fallback(cluster_texts, top_n)


def _tfidf_keywords_fallback(texts: list[str], top_n: int = 8) -> list[str]:
    from sklearn.feature_extraction.text import TfidfVectorizer
    STOP = {"i","me","my","we","our","you","your","he","she","it","they",
            "the","a","an","and","or","but","in","on","at","to","for","of",
            "with","by","from","is","are","was","were","be","been","have","has",
            "had","do","does","did","will","would","could","should","not","no",
            "this","that","what","which","who","how","when","where","there","as",
            "if","than","also","just","very","too","only","all","any","so","yet",
            "even","up","out","its","don't","can't","won't","didn't","doesn't"}
    try:
        vec = TfidfVectorizer(stop_words=list(STOP), ngram_range=(1, 2),
                              min_df=1, max_features=200, sublinear_tf=True)
        mat = vec.fit_transform([" ".join(texts)])
        feat = vec.get_feature_names_out()
        scores = mat[0].toarray()[0]
        top_idx = scores.argsort()[::-1][:top_n]
        return [feat[i] for i in top_idx if scores[i] > 0]
    except Exception:
        return []


def label_node_with_llm(node_level, keywords, sample_reviews, avg_sentiment,
                        review_count, used_themes=None):
    level_instructions = {
        "L1": ('Generate a SHORT broad area label (2-4 words). '
               'Not generic like "App Experience". '
               'Return ONLY valid JSON: {"label": "...", "description": "..."}'),
        "L2": ('Generate a FEATURE AREA label (2-5 words). '
               'Return ONLY valid JSON: {"label": "...", "description": "..."}'),
        "L3": ('Generate a FULL 5-LEVEL TAXONOMY. '
               'The "theme" field MUST be unique. Already taken themes:\n'
               '{taken_themes_placeholder}\n'
               'Return ONLY valid JSON with keys: l1, l2, l3, theme, subtheme, '
               'problem_statement, root_cause, recommended_actions (list of 3).'),
    }

    rendered = level_instructions.copy()
    if node_level == "L3" and used_themes:
        taken = "\n".join(f"  - {t}" for t in used_themes)
        rendered["L3"] = rendered["L3"].replace("{taken_themes_placeholder}", taken)
    else:
        rendered["L3"] = rendered["L3"].replace("{taken_themes_placeholder}", "  (none yet)")

    prompt = (f"You are a senior product analyst.\n\n"
              f"CLUSTER DATA:\n- Level: {node_level}\n- Reviews: {review_count}\n"
              f"- Avg sentiment: {avg_sentiment:.2f}\n"
              f"- Keywords: {', '.join(keywords[:8])}\n\n"
              f"SAMPLE REVIEWS:\n" +
              "\n".join(f'{i+1}. "{r}"' for i, r in enumerate(sample_reviews[:6])) +
              f"\n\nINSTRUCTIONS:\n{rendered[node_level]}")

    try:
        api_key = os.environ.get("ANTHROPIC_API_KEY", "")
        if not api_key:
            raise ValueError("ANTHROPIC_API_KEY not set.")
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model="claude-sonnet-4-20250514", max_tokens=600,
            messages=[{"role": "user", "content": prompt}])
        raw = response.content[0].text.strip()
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
        return json.loads(raw)
    except Exception as e:
        print(f"\n    LLM call failed for {node_level}: {e}")
        return _fallback_label(node_level, keywords)


def _fallback_label(node_level, keywords):
    base = f"Theme: {', '.join(keywords[:3])}" if keywords else "Cluster"
    if node_level in ("L1", "L2"):
        return {"label": base, "description": "Auto-labelled from keywords"}
    return {"l1": "Uncategorised", "l2": "Unknown", "l3": "Unknown",
            "theme": base, "subtheme": "LLM unavailable",
            "problem_statement": "Could not generate insight.",
            "root_cause": "LLM call failed.",
            "recommended_actions": ["Review this cluster manually"]}


def generate_l1_vocabulary(cluster_summaries, previous_vocab=None):
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key or not api_key.startswith("sk-ant-"):
        return _fallback_vocab_from_keywords(cluster_summaries)
    cluster_lines = []
    for i, c in enumerate(cluster_summaries):
        kws = c.get("keywords", [])[:6]
        sent = c.get("avg_sentiment", 0.0)
        revs = c.get("sample_reviews", [])[:2]
        cluster_lines.append(f"Cluster {i} (sentiment={sent:+.2f}): keywords=[{', '.join(kws)}] | sample=[{' / '.join(r[:70] for r in revs)}]")
    evolution_block = ""
    if previous_vocab:
        evolution_block = "\nPREVIOUS VOCABULARY:\n" + "\n".join(f"  - {v}" for v in previous_vocab) + "\n"
    prompt = ("You are a senior product analyst building a taxonomy.\n\nCLUSTER SUMMARIES:\n" +
              "\n".join(cluster_lines) + evolution_block +
              "\nGenerate 3-6 broad L1 area labels. 2-4 words, specific, not generic.\n"
              'Return ONLY valid JSON list: ["Label One", "Label Two", ...]')
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(model="claude-sonnet-4-20250514", max_tokens=200,
                                          messages=[{"role":"user","content":prompt}])
        raw = response.content[0].text.strip()
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
        vocab = json.loads(raw)
        if isinstance(vocab, list) and all(isinstance(v, str) for v in vocab):
            print(f"  Generated L1 vocabulary: {vocab}")
            return vocab
    except Exception as e:
        print(f"  L1 vocabulary generation failed: {e}")
    return _fallback_vocab_from_keywords(cluster_summaries)

def _fallback_vocab_from_keywords(cluster_summaries):
    all_kw = []
    for c in cluster_summaries:
        all_kw.extend(c.get("keywords", [])[:3])
    buckets = {
        "Monetisation": ["credit","money","billing","charge","payment","cost","price"],
        "Product Reliability": ["deploy","crash","blank","screen","bug","glitch","error"],
        "AI Quality": ["ai","agent","memory","forgetful","irrelevant","answer"],
        "Onboarding": ["free","install","signup","first","register","expo"],
        "User Experience": ["slow","loading","confusing","ui","design","interface"],
        "User Delight": ["great","love","awesome","perfect","career","vibe"],
        "Trust & Safety": ["scam","fraud","trust","security","stolen"],
    }
    found = set()
    kw_text = " ".join(all_kw).lower()
    for label, kws in buckets.items():
        if any(k in kw_text for k in kws):
            found.add(label)
    return list(found) if found else ["Product Experience", "User Feedback"]

def assign_l1_globally(cluster_summaries, previous_vocab=None):
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    vocab = generate_l1_vocabulary(cluster_summaries, previous_vocab)
    if not api_key or not api_key.startswith("sk-ant-"):
        return _fallback_l1_assignment(cluster_summaries, vocab)
    cluster_lines = []
    for i, c in enumerate(cluster_summaries):
        kws = c.get("keywords", [])[:5]
        sent = c.get("avg_sentiment", 0.0)
        revs = c.get("sample_reviews", [])[:2]
        cluster_lines.append(f"Cluster {i}: keywords=[{', '.join(kws)}] sentiment={sent:+.2f} sample=[{' / '.join(r[:60] for r in revs)}]")
    prompt = ("Classify review clusters.\n\nL1 VOCABULARY:\n" +
              "\n".join(f"  - {v}" for v in vocab) +
              "\n\nCLUSTERS:\n" + "\n".join(cluster_lines) +
              '\n\nAssign each cluster to ONE label from vocabulary.\nReturn ONLY valid JSON dict: {"0":"...","1":"...",...}')
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(model="claude-sonnet-4-20250514", max_tokens=300,
                                          messages=[{"role":"user","content":prompt}])
        raw = response.content[0].text.strip()
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
        result = json.loads(raw)
        print(f"  Global L1 assignments: {result}")
        return result
    except Exception as e:
        print(f"  Global L1 assignment failed: {e}")
    return _fallback_l1_assignment(cluster_summaries, vocab)

def _fallback_l1_assignment(cluster_summaries, vocab=None):
    if vocab:
        assignments = {}
        for i, c in enumerate(cluster_summaries):
            kw_text = " ".join(c.get("keywords", [])).lower()
            best_label, best_score = vocab[0], 0
            for label in vocab:
                score = sum(1 for w in label.lower().split() if w in kw_text)
                if score > best_score:
                    best_score = score
                    best_label = label
            assignments[str(i)] = best_label
        return assignments
    rules = [
        (["credit","money","billing","charge","cost","payment"], "Monetisation"),
        (["deploy","crash","blank","screen","bug","glitch"], "Product Reliability"),
        (["ai","agent","memory","forgetful","irrelevant"], "AI Agent Quality"),
        (["free","onboard","signup","register","expo"], "Acquisition & Onboarding"),
        (["great","love","awesome","perfect","vibe"], "User Delight"),
    ]
    assignments = {}
    for i, c in enumerate(cluster_summaries):
        kw_text = " ".join(c.get("keywords", [])).lower()
        assigned = "Product Experience"
        for kws, l1 in rules:
            if any(k in kw_text for k in kws):
                assigned = l1
                break
        assignments[str(i)] = assigned
    return assignments

def label_full_hierarchy(reviews, doc_topics, topic_tree, flat_topics, st_model):
    use_llm = bool(os.environ.get("ANTHROPIC_API_KEY", "").startswith("sk-ant-"))
    if not use_llm:
        key = os.environ.get("ANTHROPIC_API_KEY", "")
        if key and not key.startswith("sk-ant-"):
            print(f"\n  Wrong API key format: '{key[:12]}...' - Anthropic keys start with 'sk-ant-'.")
        else:
            print("  ANTHROPIC_API_KEY not set - using keyword fallback labels.")

    texts = [r["text"] for r in reviews]

    l1_docs = defaultdict(list)
    l2_docs = defaultdict(list)
    l3_docs = defaultdict(list)
    for rev_idx, topic_id in enumerate(doc_topics):
        if topic_id == -1 or topic_id not in topic_tree:
            continue
        node = topic_tree[topic_id]
        l1_docs[node["l1_id"]].append(rev_idx)
        l2_docs[node["l2_id"]].append(rev_idx)
        l3_docs[node["l3_id"]].append(rev_idx)

    node_labels = {}

    def node_data(doc_indices):
        node_texts = [texts[i] for i in doc_indices]
        node_embs = np.array([flat_topics.get(doc_topics[i], {}).get("centroid", np.zeros(384)) for i in doc_indices])
        centroid = node_embs.mean(axis=0)
        sentiments = [score_sentiment(t)["compound"] for t in node_texts]
        return node_texts, centroid, float(np.mean(sentiments))

    print("\n  Pass 1: Global L1 assignment...")
    l3_summaries = []
    l3_ids_ordered = sorted(l3_docs.keys())
    for l3_id in l3_ids_ordered:
        doc_indices = l3_docs[l3_id]
        node_texts, centroid, avg_sent = node_data(doc_indices)
        keywords = extract_node_keywords(node_texts, centroid, st_model, top_n=6)
        l3_summaries.append({"l3_id": l3_id, "keywords": keywords, "avg_sentiment": avg_sent, "sample_reviews": node_texts[:3]})

    history_for_vocab = load_taxonomy_history(TAXONOMY_HISTORY_FILE)
    previous_vocab = None
    if history_for_vocab:
        last_date = max(history_for_vocab.keys())
        prev_snap = history_for_vocab[last_date]
        previous_vocab = list({v.get("l1") for v in prev_snap.values() if v.get("l1")})

    global_l1 = assign_l1_globally(l3_summaries, previous_vocab=previous_vocab)
    l3_to_l1 = {}
    for idx, label in global_l1.items():
        if str(idx).isdigit() and int(idx) < len(l3_ids_ordered):
            l3_to_l1[l3_ids_ordered[int(idx)]] = label
    print(f"  L1 assignments: {l3_to_l1}")

    print("\n  Building L1 node labels...")
    for l1_id, doc_indices in l1_docs.items():
        node_texts, centroid, avg_sent = node_data(doc_indices)
        keywords = extract_node_keywords(node_texts, centroid, st_model, top_n=8)
        l1_label = "Product Experience"
        for l3_id, assigned_l1 in l3_to_l1.items():
            if topic_tree.get(l3_id, {}).get("l1_id") == l1_id:
                l1_label = assigned_l1
                break
        node_labels[("L1", l1_id)] = {"label": l1_label, "keywords": keywords, "doc_indices": doc_indices}
        print(f"    L1 {l1_id} ({len(doc_indices)} reviews) -> {l1_label}")

    print("\n  Labelling L2 nodes...")
    for l2_id, doc_indices in l2_docs.items():
        node_texts, centroid, avg_sent = node_data(doc_indices)
        keywords = extract_node_keywords(node_texts, centroid, st_model, top_n=8)
        if use_llm:
            label = label_node_with_llm("L2", keywords, node_texts, avg_sent, len(doc_indices))
        else:
            label = _fallback_label("L2", keywords)
        node_labels[("L2", l2_id)] = {**label, "keywords": keywords, "doc_indices": doc_indices}
        print(f"    L2 {l2_id} ({len(doc_indices)} reviews) -> {label.get('label', '?')}")

    print("\n  Labelling L3 nodes (full taxonomy)...")
    used_themes = []
    sorted_l3 = sorted(l3_docs.items(), key=lambda x: len(x[1]), reverse=True)
    for l3_id, doc_indices in sorted_l3:
        node_texts, centroid, avg_sent = node_data(doc_indices)
        keywords = extract_node_keywords(node_texts, centroid, st_model, top_n=8)
        if use_llm:
            taxonomy = label_node_with_llm("L3", keywords, node_texts, avg_sent, len(doc_indices), used_themes=used_themes)
        else:
            taxonomy = _fallback_label("L3", keywords)
        node_labels[("L3", l3_id)] = {**taxonomy, "keywords": keywords, "doc_indices": doc_indices}
        theme = taxonomy.get("theme", taxonomy.get("label", "?"))
        if theme and theme != "?":
            used_themes.append(theme)
        print(f"    L3 {l3_id} ({len(doc_indices)} reviews) -> {theme[:50]}")

    doc_taxonomy = {}
    for rev_idx, topic_id in enumerate(doc_topics):
        if topic_id == -1 or topic_id not in topic_tree:
            doc_taxonomy[rev_idx] = {"l1": "Noise", "l2": "Unclassified", "l3": "Unclassified",
                                     "theme": "Outlier / Noise", "subtheme": "Low-confidence", "topic_id": -1}
            continue
        node = topic_tree[topic_id]
        l1_info = node_labels.get(("L1", node["l1_id"]), {})
        l2_info = node_labels.get(("L2", node["l2_id"]), {})
        l3_info = node_labels.get(("L3", node["l3_id"]), {})
        doc_taxonomy[rev_idx] = {
            "l1": l1_info.get("label", l3_info.get("l1", "Unknown")),
            "l2": l2_info.get("label", l3_info.get("l2", "Unknown")),
            "l3": l3_info.get("l3", "Unknown"),
            "theme": l3_info.get("theme", "Unknown"),
            "subtheme": l3_info.get("subtheme", ""),
            "problem_statement": l3_info.get("problem_statement", ""),
            "root_cause": l3_info.get("root_cause", ""),
            "recommended_actions": l3_info.get("recommended_actions", []),
            "topic_id": topic_id,
            "l1_id": node["l1_id"], "l2_id": node["l2_id"], "l3_id": node["l3_id"],
            "l1_keywords": l1_info.get("keywords", []),
            "l2_keywords": l2_info.get("keywords", []),
            "l3_keywords": l3_info.get("keywords", []),
        }
    return node_labels, doc_taxonomy

STOP_WORDS = {
    "i","me","my","we","our","you","your","he","she","it","they","the","a","an",
    "and","or","but","in","on","at","to","for","of","with","by","from","is","are",
    "was","were","be","been","being","have","has","had","do","does","did","will",
    "would","could","should","may","might","not","no","nor","so","yet","this","that",
    "these","those","what","which","who","how","when","where","there","here","as",
    "if","then","than","also","just","even","very","too","only","up","out","all",
    "any","its","it's","i'm","don't","can't","won't","didn't","doesn't",
    "aur","hai","ke","ki","ka","ko","se","mein","jo","toh","bahut","bas","koi","kya","ab",
}

_VADER_ANALYZER = None
_VADER_FAILED = False

def _get_vader_analyzer():
    global _VADER_ANALYZER, _VADER_FAILED
    if _VADER_FAILED:
        return None
    if _VADER_ANALYZER is None:
        try:
            from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
            _VADER_ANALYZER = SentimentIntensityAnalyzer()
        except (MemoryError, OSError):
            _VADER_FAILED = True
            return None
    return _VADER_ANALYZER

def _fallback_sentiment(text):
    tl = text.lower()
    pos = sum(1 for w in ("great","love","awesome","perfect","good","nice","amazing","excellent") if w in tl)
    neg = sum(1 for w in ("bad","terrible","awful","sucks","garbage","scam","worst","hate") if w in tl)
    c = 0.0
    if pos > neg: c = min(0.5 + (pos - neg) * 0.1, 1.0)
    elif neg > pos: c = max(-0.5 - (neg - pos) * 0.1, -1.0)
    return {"compound": round(c, 4), "positive_ratio": round(max(0, c), 4) if c > 0 else 0,
            "negative_ratio": round(max(0, -c), 4) if c < 0 else 0,
            "neutral_ratio": round(1 - abs(c), 4) if abs(c) < 1 else 0,
            "label": "Positive" if c >= 0.05 else ("Negative" if c <= -0.05 else "Neutral")}

def score_sentiment(text):
    analyzer = _get_vader_analyzer()
    if analyzer is not None:
        try:
            s = analyzer.polarity_scores(text)
            c = s["compound"]
            return {"compound": round(c, 4), "positive_ratio": round(s["pos"], 4),
                    "negative_ratio": round(s["neg"], 4), "neutral_ratio": round(s["neu"], 4),
                    "label": "Positive" if c >= 0.05 else ("Negative" if c <= -0.05 else "Neutral")}
        except MemoryError:
            pass
    return _fallback_sentiment(text)

SEVERITY_SIGNALS = {
    "critical": {"keywords": ["scam","scammed","fraud","disappeared","blank screen","lost money",
                               "lost credits","lost all","charged again","$500","not deployed",
                               "all disappeared"], "weight": 1.0},
    "high":     {"keywords": ["lost","gone","nothing","doesn't work","irrelevant","misunderstands",
                               "not worth","alternatives","garbage","sucks","forgetful",
                               "no chat history","no accountability","wasted","recharge"], "weight": 0.6},
    "medium":   {"keywords": ["too fast","expensive","spendy","too much","confusing","improvement",
                               "not free","costs","slow","glitches"], "weight": 0.3},
    "low":      {"keywords": ["bad","not good","ok","average"], "weight": 0.1},
}

def score_severity(text):
    tl = text.lower()
    raw, matched = 0.0, []
    for tier, cfg in SEVERITY_SIGNALS.items():
        for kw in cfg["keywords"]:
            if kw in tl:
                raw += cfg["weight"]
                matched.append({"keyword": kw, "tier": tier, "weight": cfg["weight"]})
    norm = min(round(raw, 4), 1.0)
    label = "Critical" if norm >= 0.8 else "High" if norm >= 0.4 else "Medium" if norm >= 0.15 else "Low"
    return {"score": norm, "label": label, "matched_signals": matched}

ACTIONABILITY_SIGNALS = {
    "specific_feature": ["chat history","memory","deploy","playstore","play store","credits",
                          "billing","support","hinglish","expo","blank screen","accuracy","subscription"],
    "quantified":       [r"\$\d+", r"\d+ credits", r"\d+ times", r"\d+ day", r"100 credits"],
    "suggests_improvement": ["please improve","should be","needs to","jarurat hai","improvement"],
    "vague_noise":      [r"^(ok|nice|bad|good|awesome|sucks?|vary bad|bakwaas)\.?$"],
}

def score_actionability(text):
    tl = text.lower().strip()
    for p in ACTIONABILITY_SIGNALS["vague_noise"]:
        if re.match(p, tl):
            return {"score": 0.0, "label": "Not actionable", "signals": ["vague_single_word"]}
    score, signals = 0.0, []
    for kw in ACTIONABILITY_SIGNALS["specific_feature"]:
        if kw in tl: score += 0.25; signals.append(kw)
    for p in ACTIONABILITY_SIGNALS["quantified"]:
        if re.search(p, tl): score += 0.3; signals.append(f"quantified:{p}")
    for ph in ACTIONABILITY_SIGNALS["suggests_improvement"]:
        if ph in tl: score += 0.2; signals.append(ph)
    wc = len(text.split())
    score += 0.2 if wc > 30 else (0.1 if wc > 10 else 0)
    norm = min(round(score, 4), 1.0)
    label = "Highly actionable" if norm >= 0.6 else "Actionable" if norm >= 0.3 else "Low actionability" if norm > 0 else "Not actionable"
    return {"score": norm, "label": label, "signals": list(set(signals))}

def tokenize(text):
    text = re.sub(r"[^\w\s$]", " ", text.lower()).split()
    return [t for t in text if t not in STOP_WORDS and len(t) > 2 and not t.isdigit()]

def get_ngrams(tokens, n):
    return [tuple(tokens[i:i+n]) for i in range(len(tokens) - n + 1)]

def analyze_word_frequencies(reviews, doc_topics):
    from sklearn.feature_extraction.text import TfidfVectorizer
    cluster_texts = defaultdict(list)
    cluster_tokens = defaultdict(list)
    for r, tid in zip(reviews, doc_topics):
        cid = int(tid)
        cluster_texts[cid].append(r["text"])
        cluster_tokens[cid].append(tokenize(r["text"]))
    cids = sorted(cluster_texts.keys())
    docs = [" ".join(cluster_texts[cid]) for cid in cids]
    tfidf_scores = {}
    try:
        vec = TfidfVectorizer(stop_words=list(STOP_WORDS), ngram_range=(1,2), min_df=1, max_features=500, sublinear_tf=True)
        mat = vec.fit_transform(docs)
        feat = vec.get_feature_names_out()
        for i, cid in enumerate(cids):
            row = mat[i].toarray()[0]
            tfidf_scores[cid] = {feat[j]: round(float(row[j]), 4) for j in range(len(feat)) if row[j] > 0}
    except Exception:
        pass
    result = {}
    for cid in cids:
        tl = cluster_tokens[cid]
        all_tok = [t for tl2 in tl for t in tl2]
        raw = Counter(all_tok)
        bg_per = [set(" ".join(bg) for bg in get_ngrams(t, 2)) for t in tl]
        tg_per = [set(" ".join(tg) for tg in get_ngrams(t, 3)) for t in tl]
        all_bg = [bg for bgs in bg_per for bg in bgs]
        all_tg = [tg for tgs in tg_per for tg in tgs]
        rep = {**{p: {"count": c, "type": "bigram"} for p, c in Counter(all_bg).items() if c >= 2},
               **{p: {"count": c, "type": "trigram"} for p, c in Counter(all_tg).items() if c >= 2}}
        wcd, seen = [], set()
        for word, freq in raw.most_common(40):
            if word in seen: continue
            seen.add(word)
            wcd.append({"word": word, "raw_freq": freq, "tfidf": round(tfidf_scores.get(cid, {}).get(word, 0.0), 4),
                        "is_repeated": any(word in p for p in rep)})
        wcd.sort(key=lambda x: (-x["tfidf"], -x["raw_freq"]))
        pm = [{"term": w["word"], "raw_freq": w["raw_freq"], "tfidf": w["tfidf"],
               "review_count": sum(1 for t in tl if w["word"] in t),
               "pct_of_cluster": round(sum(1 for t in tl if w["word"] in t) / len(tl) * 100, 1)} for w in wcd[:10]]
        result[cid] = {"cluster_id": cid, "review_count": len(tl), "top_words": wcd[:20],
                       "pm_signals": pm, "repeated_phrases": rep, "raw_freq_all": dict(raw.most_common(50))}
    return result

TAXONOMY_HISTORY_FILE = "taxonomy_history.json"

def load_taxonomy_history(path=TAXONOMY_HISTORY_FILE):
    if os.path.exists(path):
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_taxonomy_snapshot(history, run_date, cluster_summary, path=TAXONOMY_HISTORY_FILE):
    snapshot = {}
    for c in cluster_summary:
        if c.get("is_noise"): continue
        cid = str(c["cluster_id"])
        kwds = sorted((c.get("l3_keywords") or [])[:3])
        sb = "positive" if c["avg_sentiment"] >= 0.05 else ("negative" if c["avg_sentiment"] <= -0.05 else "neutral")
        snapshot[cid] = {"label": c.get("cluster_label", f"Cluster {cid}"),
                         "l1": c.get("taxonomy", {}).get("l1", ""), "l2": c.get("taxonomy", {}).get("l2", ""),
                         "theme": c.get("taxonomy", {}).get("theme", ""), "top_keywords": kwds,
                         "avg_sentiment": round(c["avg_sentiment"], 4), "sentiment_bucket": sb,
                         "review_count": c["review_count"], "priority_score": c["priority_score"],
                         "fingerprint": hashlib.md5(("|".join(kwds) + sb).encode()).hexdigest()[:8]}
    history[run_date] = snapshot
    with open(path, "w", encoding="utf-8") as f:
        json.dump(history, f, indent=2, ensure_ascii=False)
    print(f"  Snapshot saved: {path} [{run_date}]")

def diff_taxonomy(history, current_snap, run_date):
    past = sorted(d for d in history if d != run_date)
    if not past:
        print("  First run - no previous snapshot.")
        return []
    prev_date = past[-1]
    prev_snap = history[prev_date]
    print(f"  Diffing vs {prev_date}")
    changes = []
    for cid, curr in current_snap.items():
        if cid in prev_snap:
            prev = prev_snap[cid]
            if curr["fingerprint"] == prev["fingerprint"]:
                changes.append({"type":"STABLE","cluster_id":cid,"label":curr["label"],"description":"No change.","severity":"LOW"})
            else:
                sf = curr["sentiment_bucket"] != prev["sentiment_bucket"]
                delta = curr["review_count"] - prev["review_count"]
                pct = round(delta / max(prev["review_count"], 1) * 100, 1)
                if sf:
                    desc = f"Sentiment: {prev['sentiment_bucket']} -> {curr['sentiment_bucket']}. Volume: {prev['review_count']} -> {curr['review_count']} ({delta:+d})."
                    sev = "HIGH"
                else:
                    desc = f"Volume: {delta:+d} ({pct:+.1f}%). Keywords: {prev['top_keywords']} -> {curr['top_keywords']}."
                    sev = "HIGH" if abs(pct) >= 30 else "MEDIUM"
                changes.append({"type":"CHANGED","cluster_id":cid,"label":curr["label"],"prev_label":prev["label"],
                                "description":desc,"severity":sev,"sentiment_delta":round(curr["avg_sentiment"]-prev["avg_sentiment"],4)})
        else:
            changes.append({"type":"NEW","cluster_id":cid,"label":curr["label"],
                            "description":f"New: \"{curr['theme']}\". {curr['review_count']} reviews.",
                            "severity":"HIGH" if curr["priority_score"]>0.4 else "MEDIUM"})
    for cid, prev in prev_snap.items():
        if cid not in current_snap:
            changes.append({"type":"RESOLVED","cluster_id":"X","label":prev["label"],
                            "description":f"\"{prev['label']}\" from {prev_date} not found.","severity":"MEDIUM"})
    changes.sort(key=lambda x:({"HIGH":0,"MEDIUM":1,"LOW":2}[x["severity"]],{"NEW":0,"CHANGED":1,"RESOLVED":2,"STABLE":3}[x["type"]]))
    return changes

def print_diff(changes):
    icons = {"NEW":"[NEW]","CHANGED":"[CHG]","RESOLVED":"[OK]","STABLE":"[ ]"}
    for c in changes:
        if c["type"] == "STABLE": continue
        print(f"  {icons[c['type']]} {c['label']}: {c['description']}")

def aggregate_clusters(reviews, doc_topics, doc_taxonomy, scored, word_freq):
    clusters = defaultdict(list)
    for i, tid in enumerate(doc_topics):
        clusters[int(tid)].append(i)
    summary = []
    for cid, indices in clusters.items():
        cr = [scored[i] for i in indices]
        tax = doc_taxonomy.get(indices[0], {}) if indices else {}
        sents = [r["sentiment"]["compound"] for r in cr]
        sevs = [r["severity"]["score"] for r in cr]
        acts = [r["actionability"]["score"] for r in cr]
        neg = sum(1 for r in cr if r["sentiment"]["label"] == "Negative")
        pos = sum(1 for r in cr if r["sentiment"]["label"] == "Positive")
        hsev = sum(1 for r in cr if r["severity"]["label"] in ("Critical", "High"))
        freq = word_freq.get(cid, {})
        summary.append({
            "cluster_id": cid, "is_noise": cid == -1,
            "cluster_label": tax.get("theme", f"Cluster {cid}"),
            "taxonomy": {"l1": tax.get("l1",""), "l2": tax.get("l2",""), "l3": tax.get("l3",""),
                         "theme": tax.get("theme",""), "subtheme": tax.get("subtheme","")},
            "problem_statement": tax.get("problem_statement", ""),
            "root_cause": tax.get("root_cause", ""),
            "recommended_actions": tax.get("recommended_actions", []),
            "l1_keywords": tax.get("l1_keywords", []),
            "l2_keywords": tax.get("l2_keywords", []),
            "l3_keywords": tax.get("l3_keywords", []),
            "review_count": len(indices),
            "review_ids": [reviews[i]["id"] for i in indices],
            "avg_sentiment": round(float(np.mean(sents)), 4),
            "avg_severity_score": round(float(np.mean(sevs)), 4),
            "avg_actionability": round(float(np.mean(acts)), 4),
            "negative_review_count": neg, "positive_review_count": pos,
            "high_severity_count": hsev,
            "priority_score": round(0.5*float(np.mean(sevs)) + 0.3*(neg/len(indices)) + 0.2*min(len(indices)/10,1.0), 4),
            "pm_signals": freq.get("pm_signals", []),
            "repeated_phrases": freq.get("repeated_phrases", {}),
            "word_cloud_data": freq.get("top_words", []),
        })
    summary.sort(key=lambda x: x["priority_score"], reverse=True)
    seen_themes = {}
    deduped = []
    for c in summary:
        theme = c.get("cluster_label", "")
        if theme in seen_themes:
            existing = deduped[seen_themes[theme]]
            n1, n2 = existing["review_count"], c["review_count"]
            total = n1 + n2
            existing["review_count"] = total
            existing["review_ids"] = existing["review_ids"] + c["review_ids"]
            existing["negative_review_count"] += c["negative_review_count"]
            existing["positive_review_count"] += c["positive_review_count"]
            existing["high_severity_count"] += c["high_severity_count"]
            for m in ("avg_sentiment", "avg_severity_score", "avg_actionability"):
                existing[m] = round((existing[m]*n1 + c.get(m,0)*n2)/total, 4)
            existing["priority_score"] = round(0.5*existing["avg_severity_score"] + 0.3*(existing["negative_review_count"]/total) + 0.2*min(total/10,1.0), 4)
            print(f"  Merged duplicate theme '{theme}' - {total} reviews")
        else:
            seen_themes[theme] = len(deduped)
            deduped.append(c)
    deduped.sort(key=lambda x: x["priority_score"], reverse=True)
    return deduped

def write_output_excel(output, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    HDR_FILL = PatternFill("solid", start_color="1A1F35", end_color="1A1F35")
    HDR_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    BODY_FONT = Font(name="Calibri", size=9)
    WRAP = Alignment(wrap_text=True, vertical="top")
    CENTER = Alignment(horizontal="center", vertical="center")
    THIN = Border(left=Side(style="thin", color="2D3555"), right=Side(style="thin", color="2D3555"),
                  top=Side(style="thin", color="2D3555"), bottom=Side(style="thin", color="2D3555"))
    SENT_POS = PatternFill("solid", start_color="1A3D2B", end_color="1A3D2B")
    SENT_NEG = PatternFill("solid", start_color="3D1A1A", end_color="3D1A1A")
    SENT_NEU = PatternFill("solid", start_color="2D2D1A", end_color="2D2D1A")
    SEV_CRIT = PatternFill("solid", start_color="5C0A0A", end_color="5C0A0A")
    SEV_HIGH = PatternFill("solid", start_color="5C3A0A", end_color="5C3A0A")
    SEV_MED = PatternFill("solid", start_color="5C5C0A", end_color="5C5C0A")

    def hdr(ws, row, col, val, width=None):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CENTER; c.border = THIN
        if width: ws.column_dimensions[get_column_letter(col)].width = width
    def cell(ws, row, col, val, fill=None, font=None, align=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = font or BODY_FONT; c.alignment = align or WRAP; c.border = THIN
        if fill: c.fill = fill

    reviews = output.get("reviews", [])
    clusters = output.get("cluster_summary", [])
    meta = output.get("meta", {})

    ws1 = wb.active; ws1.title = "Reviews"
    ws1.merge_cells("A1:N1")
    t = ws1["A1"]
    t.value = f"Review Analysis - {meta.get('run_date','')} | {len(reviews)} reviews | {meta.get('n_clusters',0)} clusters"
    t.fill = HDR_FILL; t.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11); t.alignment = CENTER
    cols = [("ID",4),("Name",18),("Date",12),("Review",55),("Sentiment",11),("Sent Score",10),
            ("Severity",10),("Sev Score",9),("Actionability",12),("Act Score",9),("Cluster",22),("L1",18),("L2",18),("Confidence",11)]
    for ci, (name, width) in enumerate(cols, 1):
        hdr(ws1, 2, ci, name, width)
    ws1.freeze_panes = "A3"
    for ri, r in enumerate(reviews, 3):
        sent = r.get("sentiment",{}); sev = r.get("severity",{}); act = r.get("actionability",{})
        tax = r.get("taxonomy",{}); sl = sent.get("label",""); svl = sev.get("label","")
        sf = SENT_POS if sl=="Positive" else (SENT_NEG if sl=="Negative" else SENT_NEU)
        svf = SEV_CRIT if svl=="Critical" else (SEV_HIGH if svl=="High" else (SEV_MED if svl=="Medium" else None))
        cell(ws1,ri,1,r.get("id","")); cell(ws1,ri,2,r.get("name","")); cell(ws1,ri,3,r.get("date",""))
        cell(ws1,ri,4,r.get("text","")); cell(ws1,ri,5,sl,fill=sf); cell(ws1,ri,6,sent.get("compound",""))
        cell(ws1,ri,7,svl,fill=svf); cell(ws1,ri,8,round(sev.get("score",0)*100,1))
        cell(ws1,ri,9,act.get("label","")); cell(ws1,ri,10,round(act.get("score",0)*100,1))
        cell(ws1,ri,11,r.get("cluster_label","")); cell(ws1,ri,12,tax.get("l1",""))
        cell(ws1,ri,13,tax.get("l2","")); cell(ws1,ri,14,round(r.get("cluster_confidence",0)*100,1))

    ws2 = wb.create_sheet("Cluster Summary")
    cols2 = [("Theme",28),("Reviews",8),("Priority",13),("Sentiment",13),("Severity",16),
             ("High Sev",13),("Actionability",20),("Keywords",30),("Problem",50),("Actions",55)]
    for ci, (name, width) in enumerate(cols2, 1):
        hdr(ws2, 1, ci, name, width)
    ws2.freeze_panes = "A2"
    for ri, c in enumerate([x for x in clusters if not x.get("is_noise")], 2):
        priority = round(c.get("priority_score",0)*100, 1)
        cell(ws2,ri,1,c.get("cluster_label","")); cell(ws2,ri,2,c.get("review_count",0))
        cell(ws2,ri,3,priority); cell(ws2,ri,4,round(c.get("avg_sentiment",0),3))
        cell(ws2,ri,5,round(c.get("avg_severity_score",0)*100,1)); cell(ws2,ri,6,c.get("high_severity_count",0))
        cell(ws2,ri,7,round(c.get("avg_actionability",0)*100,1))
        cell(ws2,ri,8,", ".join((c.get("l3_keywords") or c.get("l1_keywords") or [])[:5]))
        cell(ws2,ri,9,c.get("problem_statement","")); actions = c.get("recommended_actions",[])
        cell(ws2,ri,10,"\n".join(f"{i+1}. {a}" for i,a in enumerate(actions[:4])))

    ws3 = wb.create_sheet("Taxonomy Tree")
    cols3 = [("L1",22),("L2",22),("L3",22),("Theme",28),("Subtheme",40),("Reviews",8),("Priority",13)]
    for ci, (name, width) in enumerate(cols3, 1):
        hdr(ws3, 1, ci, name, width)
    ws3.freeze_panes = "A2"
    non_noise = sorted([c for c in clusters if not c.get("is_noise")],
                       key=lambda x: (x.get("taxonomy",{}).get("l1",""), x.get("taxonomy",{}).get("l2","")))
    for ri, c in enumerate(non_noise, 2):
        tax = c.get("taxonomy",{})
        cell(ws3,ri,1,tax.get("l1","")); cell(ws3,ri,2,tax.get("l2","")); cell(ws3,ri,3,tax.get("l3",""))
        cell(ws3,ri,4,tax.get("theme","")); cell(ws3,ri,5,tax.get("subtheme",""))
        cell(ws3,ri,6,c.get("review_count",0)); cell(ws3,ri,7,round(c.get("priority_score",0)*100,1))

    ws4 = wb.create_sheet("CX Metrics")
    hdr(ws4,1,1,"Metric",28); hdr(ws4,1,2,"Value",14); hdr(ws4,1,3,"Formula",40); hdr(ws4,1,4,"Interpretation",50)
    pos = sum(1 for r in reviews if r.get("sentiment",{}).get("label")=="Positive")
    neg = sum(1 for r in reviews if r.get("sentiment",{}).get("label")=="Negative")
    total = max(len(reviews), 1)
    promoters = sum(1 for r in reviews if r.get("sentiment",{}).get("compound",0)>=0.5 and r.get("severity",{}).get("score",0)<0.2)
    detractors = sum(1 for r in reviews if r.get("sentiment",{}).get("compound",0)<=-0.3 or r.get("severity",{}).get("label") in ("Critical","High"))
    csat = round(pos/total*100, 1)
    nps = round((promoters-detractors)/total*100, 1)
    metrics = [
        ("CSAT", f"{csat}%", "% Positive reviews", "Good" if csat>=50 else "Needs attention"),
        ("NPS Proxy", f"{nps:+.1f}", "(promoters-detractors)/total*100", "Positive" if nps>0 else "Negative"),
        ("Promoters", promoters, "sentiment>=0.5 AND severity<0.2", f"{round(promoters/total*100,1)}%"),
        ("Detractors", detractors, "sentiment<=-0.3 OR severity Critical/High", f"{round(detractors/total*100,1)}%"),
        ("Total Reviews", total, "", ""),
        ("Negative Reviews", neg, "", f"{round(neg/total*100,1)}%"),
        ("High Severity", sum(1 for r in reviews if r.get("severity",{}).get("label") in ("Critical","High")), "", ""),
    ]
    for ri, (m, v, f, interp) in enumerate(metrics, 2):
        cell(ws4,ri,1,m,font=Font(name="Calibri",size=9,bold=True)); cell(ws4,ri,2,v,align=CENTER)
        cell(ws4,ri,3,f); cell(ws4,ri,4,interp)

    wb.save(path)
    print(f"  Excel output -> {path}")

def run_pipeline(reviews=None, output_path="review_analysis.json",
                 history_path=TAXONOMY_HISTORY_FILE, run_date=None):
    if reviews is None: reviews = RAW_REVIEWS
    if run_date is None: run_date = datetime.today().strftime("%Y-%m-%d")
    print("\n" + "="*65)
    print("  VOICE OF CUSTOMER INTELLIGENCE - Pipeline v4")
    print("  Method : BERTopic Hierarchy + KeyBERT + LLM Labels")
    print(f"  Reviews: {len(reviews)}  |  Date: {run_date}")
    print("="*65 + "\n")
    texts = [r["text"] for r in reviews]

    print("-- Step 1: Scoring reviews --")
    scored = []
    for r in reviews:
        scored.append({**r, "sentiment": score_sentiment(r["text"]),
                       "severity": score_severity(r["text"]), "actionability": score_actionability(r["text"]),
                       "cluster_id": None, "cluster_label": None, "cluster_confidence": None,
                       "is_noise": None, "is_representative": False, "umap_2d_x": None, "umap_2d_y": None, "taxonomy": {}})
    print(f"  Scored {len(scored)} reviews.\n")

    print("-- Step 2: Embedding --")
    embeddings, st_model = embed_reviews(texts)

    print("\n-- Step 3: BERTopic hierarchical clustering --")
    topic_model, doc_topics, hierarchy_df, topic_tree, flat_topics = build_topic_hierarchy(texts, embeddings)

    print("\n-- Step 4: Word frequency analysis --")
    word_freq = analyze_word_frequencies(reviews, doc_topics)

    print("\n-- Step 5: LLM taxonomy labelling --")
    node_labels, doc_taxonomy = label_full_hierarchy(reviews, doc_topics, topic_tree, flat_topics, st_model)

    print("\n-- Step 6: 2D UMAP --")
    from umap import UMAP
    coords_2d = UMAP(n_components=2, n_neighbors=5, min_dist=0.15, metric="cosine", random_state=42).fit_transform(embeddings)

    for i, (r, tid) in enumerate(zip(scored, doc_topics)):
        cid = int(tid)
        tax = doc_taxonomy.get(i, {})
        r["cluster_id"] = cid
        r["cluster_label"] = tax.get("theme", f"Cluster {cid}")
        r["is_noise"] = cid == -1
        r["is_representative"] = (flat_topics.get(cid, {}).get("docs", [None])[0] == i)
        dist = np.linalg.norm(embeddings[i] - flat_topics.get(cid, {}).get("centroid", embeddings[i]))
        r["cluster_confidence"] = float(1.0 - dist / max(1e-6, dist + 1)) if cid != -1 else 0.0
        r["umap_2d_x"] = round(float(coords_2d[i][0]), 6)
        r["umap_2d_y"] = round(float(coords_2d[i][1]), 6)
        r["taxonomy"] = {"l1": tax.get("l1",""), "l2": tax.get("l2",""), "l3": tax.get("l3",""),
                         "theme": tax.get("theme",""), "subtheme": tax.get("subtheme","")}

    print("\n-- Step 7: Aggregating cluster summaries --")
    cluster_summary = aggregate_clusters(reviews, doc_topics, doc_taxonomy, scored, word_freq)

    print("\n-- Step 8: Taxonomy diff --")
    history = load_taxonomy_history(history_path)
    current_snap = {}
    for c in cluster_summary:
        if not c["is_noise"]:
            cid = str(c["cluster_id"])
            kwds = sorted((c.get("l3_keywords") or c.get("l1_keywords") or [])[:3])
            sb = "positive" if c["avg_sentiment"] >= 0.05 else ("negative" if c["avg_sentiment"] <= -0.05 else "neutral")
            current_snap[cid] = {"label": c["cluster_label"], "l1": c["taxonomy"].get("l1",""),
                                 "l2": c["taxonomy"].get("l2",""), "theme": c["taxonomy"].get("theme",""),
                                 "top_keywords": kwds, "avg_sentiment": c["avg_sentiment"], "sentiment_bucket": sb,
                                 "review_count": c["review_count"], "priority_score": c["priority_score"],
                                 "fingerprint": hashlib.md5(("|".join(kwds)+sb).encode()).hexdigest()[:8]}
    diff_changes = diff_taxonomy(history, current_snap, run_date)
    if diff_changes: print_diff(diff_changes)
    save_taxonomy_snapshot(history, run_date, cluster_summary, history_path)

    output = {
        "meta": {"run_date": run_date, "total_reviews": len(reviews),
                 "n_clusters": len([c for c in cluster_summary if not c["is_noise"]]),
                 "n_noise_points": sum(1 for t in doc_topics if t == -1),
                 "pipeline_version": "v4",
                 "methods": {"embedding": "sentence-transformers/all-MiniLM-L6-v2",
                             "clustering": "BERTopic (UMAP + HDBSCAN + Ward linkage)",
                             "keywords": "KeyBERT", "word_freq": "TF-IDF + Counter",
                             "taxonomy": "LLM (claude-sonnet-4-20250514)" if os.environ.get("ANTHROPIC_API_KEY") else "TF-IDF fallback"}},
        "taxonomy_diff": diff_changes, "cluster_summary": cluster_summary, "reviews": scored,
    }
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
    print(f"\n  JSON output -> {output_path}")

    excel_path = output_path.replace(".json", "_output.xlsx")
    if not excel_path.endswith(".xlsx"): excel_path += "_output.xlsx"
    print("-- Step 9: Writing Excel output --")
    try:
        write_output_excel(output, excel_path)
    except Exception as e:
        print(f"  Excel write failed: {e}")

    print("\n" + "="*80)
    print(f"  {'#':<3} {'L1':<20} {'L2':<22} {'Theme':<25} {'N':<4} {'P-Score'}")
    print("-"*80)
    for rank, c in enumerate(cluster_summary, 1):
        if not c["is_noise"]:
            tax = c.get("taxonomy", {})
            print(f"  {rank:<3} {tax.get('l1','?')[:19]:<20} {tax.get('l2','?')[:21]:<22} "
                  f"{tax.get('theme','?')[:24]:<25} {c['review_count']:<4} {c['priority_score']:.3f}")
    return output


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="VoC Intelligence Pipeline v4")
    parser.add_argument("--input", type=str, default=None, help=".xlsx file with reviews")
    parser.add_argument("--output", type=str, default="review_analysis.json", help="JSON output path")
    parser.add_argument("--history", type=str, default=TAXONOMY_HISTORY_FILE)
    parser.add_argument("--date", type=str, default=None, help="Override run date YYYY-MM-DD")
    args = parser.parse_args()
    if args.input:
        print(f"\n  Input  : {args.input}")
        reviews = load_from_excel(args.input)
    else:
        print("\nNo --input provided. Using built-in demo data.")
        reviews = RAW_REVIEWS
    print(f"  Output : {args.output}")
    print(f"  Excel  : {args.output.replace('.json', '_output.xlsx')}")
    run_pipeline(reviews=reviews, output_path=args.output, history_path=args.history, run_date=args.date)
